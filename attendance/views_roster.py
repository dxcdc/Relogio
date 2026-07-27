from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta
import json
from .models import ShiftOverride, WorkSchedule
from pim.models import Employee

@login_required
def roster_builder(request):
    """
    Renderiza o Painel Visual 'Escala Viva' (Roster Planner)
    """
    if not (request.user.is_admin() or request.user.is_hr() or request.user.role == 'Supervisor'):
        return render(request, 'core/403.html')
        
    
    shift_blocks = WorkSchedule.objects.filter(is_active=True).prefetch_related('days')
    
    
    
    if request.user.is_admin() or request.user.is_hr():
        employees = Employee.objects.all().order_by('first_name')
    else:
        emp = getattr(request.user, 'employee', None)
        if emp:
            if emp.sub_division:
                employees = Employee.objects.filter(sub_division=emp.sub_division).order_by('first_name')
            elif emp.job_title:
                employees = Employee.objects.filter(job_title=emp.job_title).order_by('first_name')
            else:
                employees = Employee.objects.none()
        else:
            employees = Employee.objects.none()

    return render(request, 'attendance/roster_builder.html', {
        'shift_blocks': shift_blocks,
        'employees': employees
    })

@login_required
def api_roster_fetch(request):
    """
    Retorna JSON com as escalas (ShiftOverrides) de uma semana específica
    """
    start_date_str = request.GET.get('start_date')
    department = request.GET.get('department')
    if not start_date_str:
        return JsonResponse({'error': 'start_date required'}, status=400)
        
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)
        
    end_date = start_date + timedelta(days=6)
    
    
    overrides_qs = ShiftOverride.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('employee')
    
    if department and department != 'all':
        overrides_qs = overrides_qs.filter(employee__sub_division__name=department)
    
    if not (request.user.is_admin() or request.user.is_hr()):
        emp = getattr(request.user, 'employee', None)
        if emp:
            if emp.sub_division:
                overrides_qs = overrides_qs.filter(employee__sub_division=emp.sub_division)
            elif emp.job_title:
                overrides_qs = overrides_qs.filter(employee__job_title=emp.job_title)
            else:
                overrides_qs = overrides_qs.none()
        else:
            overrides_qs = overrides_qs.none()
            
    data = []
    for ov in overrides_qs:
        data.append({
            'id': ov.id,
            'employee_id': ov.employee.id,
            'employee_name': ov.employee.full_name,
            'date': ov.date.isoformat(),
            'entry_time': ov.entry_time.strftime('%H:%M') if ov.entry_time else None,
            'exit_time': ov.exit_time.strftime('%H:%M') if ov.exit_time else None,
            'type': ov.override_type
        })
        
    return JsonResponse({'overrides': data})

@login_required
def api_roster_update(request):
    """
    Recebe um payload JSON pra criar/deletar blocos de ShiftOverride.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)
        
    try:
        payload = json.loads(request.body)
        action = payload.get('action') 
        
        if action == 'add':
            emp_id = payload.get('employee_id')
            date_str = payload.get('date')
            block_id = payload.get('block_id') 
            
            if not all([emp_id, date_str, block_id]):
                return JsonResponse({'error': 'Missing parameters'}, status=400)
                
            emp = Employee.objects.get(pk=emp_id)
            
            
            if not (request.user.is_admin() or request.user.is_hr()):
                user_emp = getattr(request.user, 'employee', None)
                if not user_emp:
                    return JsonResponse({'error': 'Você não tem perfil vinculado para escalar funcionários.'}, status=403)
                if user_emp.sub_division:
                    if emp.sub_division != user_emp.sub_division:
                        return JsonResponse({'error': 'Você não tem permissão para escalar este funcionário (Setor Restrito).'}, status=403)
                elif user_emp.job_title:
                    if emp.job_title != user_emp.job_title:
                        return JsonResponse({'error': 'Você não tem permissão para escalar este funcionário (Cargo Restrito).'}, status=403)
                else:
                    return JsonResponse({'error': 'Você não possui setor associado para escalar funcionários.'}, status=403)
                    
            block = WorkSchedule.objects.get(pk=block_id)
            date_val = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            
            day_config = block.days.filter(weekday=date_val.weekday()).first()
            target_entry = day_config.entry_time if day_config and day_config.is_work_day else block.display_entry
            target_exit = day_config.exit_time if day_config and day_config.is_work_day else block.display_exit
            
            if not target_entry or not target_exit:
                 return JsonResponse({'error': 'Este turno não possui expediente configurado para esse dia da semana.'}, status=400)
            
            
            ShiftOverride.objects.filter(employee=emp, date=date_val).delete()
            
            ov = ShiftOverride.objects.create(
                employee=emp,
                date=date_val,
                override_type=ShiftOverride.TYPE_WORK,
                entry_time=target_entry,
                exit_time=target_exit,
                reason=f"Escala Viva - {block.name}",
                created_by=request.user
            )
            return JsonResponse({'success': True, 'override_id': ov.id})
            
        elif action == 'remove':
            override_id = payload.get('override_id')
            if not override_id:
                return JsonResponse({'error': 'Missing override_id'}, status=400)
                
            try:
                override = ShiftOverride.objects.get(pk=override_id)
                
                if not (request.user.is_admin() or request.user.is_hr()):
                    user_emp = getattr(request.user, 'employee', None)
                    if not user_emp:
                        return JsonResponse({'error': 'Você não tem perfil vinculado para remover turnos.'}, status=403)
                    
                    if user_emp.sub_division:
                        if override.employee.sub_division != user_emp.sub_division:
                            return JsonResponse({'error': 'Você não tem permissão para remover o turno deste funcionário (Setor Restrito).'}, status=403)
                    elif user_emp.job_title:
                        if override.employee.job_title != user_emp.job_title:
                            return JsonResponse({'error': 'Você não tem permissão para remover o turno deste funcionário (Cargo Restrito).'}, status=403)
                    else:
                        return JsonResponse({'error': 'Você não possui setor associado para remover turnos.'}, status=403)
                        
                override.delete()
                return JsonResponse({'success': True})
            except ShiftOverride.DoesNotExist:
                return JsonResponse({'error': 'Turno não encontrado.'}, status=404)
            
        elif action in ['clear_week', 'clear_month']:
            department = payload.get('department')
            
            if action == 'clear_week':
                date_str = payload.get('date')
                if not date_str:
                    return JsonResponse({'error': 'Missing date'}, status=400)
                start_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                end_date = start_date + timedelta(days=6)
                overrides_qs = ShiftOverride.objects.filter(date__range=[start_date, end_date])
            else:
                month_str = payload.get('month')
                if not month_str:
                    return JsonResponse({'error': 'Missing month'}, status=400)
                
                import calendar
                year, month = map(int, month_str.split('-'))
                start_date = datetime(year, month, 1).date()
                last_day = calendar.monthrange(year, month)[1]
                end_date = datetime(year, month, last_day).date()
                overrides_qs = ShiftOverride.objects.filter(date__range=[start_date, end_date])
            
            if department and department != 'all':
                overrides_qs = overrides_qs.filter(employee__sub_division__name=department)
            
            
            if not (request.user.is_admin() or request.user.is_hr()):
                user_emp = getattr(request.user, 'employee', None)
                if not user_emp:
                    return JsonResponse({'error': 'Você não tem perfil vinculado para remover turnos.'}, status=403)
                if user_emp.sub_division:
                    overrides_qs = overrides_qs.filter(employee__sub_division=user_emp.sub_division)
                elif user_emp.job_title:
                    overrides_qs = overrides_qs.filter(employee__job_title=user_emp.job_title)
                else:
                    overrides_qs = overrides_qs.none()
                    
            count = overrides_qs.count()
            overrides_qs.delete()
            return JsonResponse({'success': True, 'message': f'{count} turnos removidos.'})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
        
    return JsonResponse({'error': 'Unknown action'}, status=400)



import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import calendar

@login_required
def api_export_roster_excel(request):
    """
    Exporta a planilha de escala para o mês selecionado
    """
    if not (request.user.is_admin() or request.user.is_hr() or request.user.role == 'Supervisor'):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
        
    start_date_str = request.GET.get('start_date')
    department = request.GET.get('department')
    if not start_date_str:
        return JsonResponse({'error': 'start_date required'}, status=400)
        
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    if request.user.is_admin() or request.user.is_hr():
        employees = Employee.objects.filter(state='ACTIVE')
        if department and department != 'all':
            employees = employees.filter(sub_division__name=department)
        employees = employees.order_by('first_name')
    else:
        emp = getattr(request.user, 'employee', None)
        if emp and emp.sub_division:
            employees = Employee.objects.filter(sub_division=emp.sub_division, state='ACTIVE').order_by('first_name')
        else:
            employees = Employee.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Escala Mensal"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    weekend_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")

    year = start_date.year
    month = start_date.month
    ndays = calendar.monthrange(year, month)[1]

    
    
    ws.cell(row=2, column=1, value="ID").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=2, value="Empregado").font = header_font
    ws.cell(row=2, column=2).fill = header_fill

    dates = []
    weekdays_pt = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
    
    for d in range(1, ndays + 1):
        dt = datetime(year, month, d).date()
        dates.append(dt)
        col = d + 2
        
        
        ws.cell(row=1, column=col, value=dt.strftime('%Y-%m-%d'))
        
        
        wd = weekdays_pt[dt.weekday()]
        header_text = f"{d:02d}{wd}"
        
        c = ws.cell(row=2, column=col, value=header_text)
        c.font = header_font
        c.alignment = align_center
        c.fill = header_fill

    ws.row_dimensions[1].hidden = True
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    for i in range(3, ndays + 3):
        ws.column_dimensions[get_column_letter(i)].width = 12

    overrides = ShiftOverride.objects.filter(
        date__range=[dates[0], dates[-1]]
    ).select_related('employee')
    
    ov_map = {}
    for ov in overrides:
        if ov.employee_id not in ov_map:
            ov_map[ov.employee_id] = {}
        ov_map[ov.employee_id][ov.date] = ov

    row_num = 3
    for emp in employees:
        ws.cell(row=row_num, column=1, value=emp.id)
        ws.cell(row=row_num, column=2, value=f"{emp.first_name} {emp.last_name}")
        
        for col_idx, d in enumerate(dates):
            cell = ws.cell(row=row_num, column=col_idx + 3)
            cell.alignment = align_center
            
            if d.weekday() in [5, 6]:
                cell.fill = weekend_fill
                
            ov = ov_map.get(emp.id, {}).get(d)
            if ov:
                if ov.override_type == 'REST':
                    cell.value = "DSR"
                elif ov.reason and "Escala Viva" in ov.reason:
                    
                    cell.value = ov.reason.replace("Escala Viva - ", "")
                elif ov.entry_time and ov.exit_time:
                    cell.value = f"{ov.entry_time.strftime('%H:%M')} - {ov.exit_time.strftime('%H:%M')}"
                else:
                    cell.value = "Turno"
            else:
                cell.value = ""
        row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=escala_{year}_{month:02d}.xlsx'
    wb.save(response)
    return response

@login_required
def api_import_roster_excel(request):
    """
    Importa a planilha de escala.
    """
    import re
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
        
    if not (request.user.is_admin() or request.user.is_hr() or request.user.role == 'Supervisor'):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
        
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file uploaded'}, status=400)
        
    excel_file = request.FILES['file']
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active

    except Exception as e:
        return JsonResponse({'error': f'Invalid Excel file: {str(e)}'}, status=400)
        
    current_date_str = request.POST.get('current_date')
    if current_date_str:
        ref_date = datetime.strptime(current_date_str, '%Y-%m-%d').date()
    else:
        ref_date = datetime.now().date()
        
    year = ref_date.year
    month = ref_date.month

    dates = []
    max_col = ws.max_column
    
    
    start_col = 2
    for c in range(2, 6):
        val1 = ws.cell(row=1, column=c).value
        val2 = ws.cell(row=2, column=c).value
        if isinstance(val1, datetime) or isinstance(val2, datetime) or re.search(r'\d+', str(val1 or '')) or re.search(r'\d+', str(val2 or '')):
            start_col = c
            break

    for col_idx in range(start_col, max_col + 1):
        
        dt_val = ws.cell(row=1, column=col_idx).value
        
        if not dt_val:
            dt_val = ws.cell(row=2, column=col_idx).value
            
        if not dt_val:
            break
            
        if isinstance(dt_val, datetime):
            dates.append(dt_val.date())
        else:
            dt_str = str(dt_val).strip()
            
            
            parsed_full = False
            dt_str_clean = dt_str.split(' ')[0] if ' ' in dt_str else dt_str
            try:
                dates.append(datetime.strptime(dt_str_clean, '%Y-%m-%d').date())
                parsed_full = True
            except ValueError:
                try:
                    dates.append(datetime.strptime(dt_str_clean, '%d/%m/%Y').date())
                    parsed_full = True
                except ValueError:
                    pass
            
            
            if not parsed_full:
                match = re.search(r'\d+', dt_str)
                if match:
                    day = int(match.group())
                    if dates and day <= dates[-1].day:
                        break 
                    try:
                        dates.append(datetime(year, month, day).date())
                    except ValueError:
                        break 
                else:
                    break 
            
    if not dates:
        return JsonResponse({'error': f'Planilha inválida. O sistema não conseguiu encontrar datas ou dias válidos (ex: "01", "02 Sáb") na linha 1 ou 2.'}, status=400)
        
    
    
    min_data_row = 3
    for r in range(2, min(10, ws.max_row + 1)):
        val0 = str(ws.cell(row=r, column=1).value or '').strip()
        val1 = str(ws.cell(row=r, column=2).value or '').strip()
        if (val0 and val0.upper() not in ['ID', 'EMPREGADO']) or (val1 and val1.upper() not in ['ID', 'EMPREGADO']):
            min_data_row = r
            break

    from attendance.models import WorkSchedule
    shift_patterns = list(WorkSchedule.objects.all())
    
    updates = 0
    from django.db import transaction
    
    def parse_time(t_str):
        
        t_str = str(t_str).lower().replace('hs', '').replace('h', '').strip()
        if ':' in t_str:
            return datetime.strptime(t_str, '%H:%M').time()
        else:
            return datetime.strptime(t_str, '%H').time()
            
    def normalize_str(s):
        import unicodedata
        if not s: return ""
        
        return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn').strip().lower()
    
    active_emps = list(Employee.objects.filter(state='ACTIVE'))
    
    with transaction.atomic():
        for row in ws.iter_rows(min_row=min_data_row, values_only=True):
            val0 = str(row[0]).strip() if row[0] else ""
            val1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            
            if not val0 and not val1:
                continue
                
            emp = None
            if val0.isdigit():
                emp = next((e for e in active_emps if e.id == int(val0)), None)
                
            if not emp and val0:
                
                search_val = normalize_str(val0)
                emp = next((e for e in active_emps if normalize_str(e.full_name) == search_val or normalize_str(f"{e.first_name} {e.last_name}") == search_val), None)
                if not emp:
                    emp = next((e for e in active_emps if normalize_str(e.full_name).startswith(search_val) or normalize_str(f"{e.first_name} {e.last_name}").startswith(search_val)), None)
                    
            if not emp and val1:
                
                search_val = normalize_str(val1)
                emp = next((e for e in active_emps if normalize_str(e.full_name) == search_val or normalize_str(f"{e.first_name} {e.last_name}") == search_val), None)
                if not emp:
                    emp = next((e for e in active_emps if normalize_str(e.full_name).startswith(search_val) or normalize_str(f"{e.first_name} {e.last_name}").startswith(search_val)), None)
            
            with open("debug_excel.txt", "a", encoding="utf-8") as f:
                f.write(f"val0='{val0}', val1='{val1}', emp={emp.full_name if emp else 'NONE'}\n")
            
            if not emp:
                continue
                
            for col_idx, d in enumerate(dates):
                if col_idx + start_col - 1 >= len(row): break
                val = row[col_idx + start_col - 1]
                
                if val is None or str(val).strip() == "":
                    ShiftOverride.objects.filter(employee=emp, date=d).delete()
                    updates += 1
                else:
                    
                    from datetime import time
                    if isinstance(val, datetime):
                        val_str = f"{val.month:02d}-{val.day:02d}"
                    elif isinstance(val, time):
                        val_str = val.strftime('%H:%M')
                    else:
                        val_str = str(val).strip()
                        
                    if not val_str or val_str.lower() in ['none', 'null', 'na', 'n/a']:
                        continue
                        
                    
                    val_str = val_str.replace('–', '-').replace('—', '-').replace('−', '-').replace('/', '-')
                        
                    val_upper = val_str.upper()
                    if val_upper in ['DSR', 'FOLGA', 'F', 'FERIADO']:
                        ShiftOverride.objects.update_or_create(
                            employee=emp, date=d,
                            defaults={
                                'override_type': ShiftOverride.TYPE_REST,
                                'entry_time': None, 
                                'exit_time': None,
                                'reason': 'Folga via Excel'
                            }
                        )
                        updates += 1
                    else:
                        matched_shift = next((sp for sp in shift_patterns if str(sp.name).upper() == val_str), None)
                        if matched_shift:
                            ShiftOverride.objects.update_or_create(
                                employee=emp, date=d,
                                defaults={
                                    'override_type': ShiftOverride.TYPE_WORK,
                                    'entry_time': getattr(matched_shift, 'entry_time', None), 
                                    'exit_time': getattr(matched_shift, 'exit_time', None),
                                    'reason': str(matched_shift.name)
                                }
                            )
                            updates += 1
                        else:
                            
                            if '-' in val_str:
                                parts = val_str.split('-')
                                if len(parts) == 2:
                                    try:
                                        st = parse_time(parts[0].strip())
                                        et = parse_time(parts[1].strip())
                                        ShiftOverride.objects.update_or_create(
                                            employee=emp, date=d,
                                            defaults={
                                                'override_type': ShiftOverride.TYPE_WORK,
                                                'entry_time': st, 
                                                'exit_time': et,
                                                'reason': 'Turno via Excel'
                                            }
                                        )
                                        updates += 1
                                    except ValueError:
                                        pass
                                        
    return JsonResponse({'success': True, 'message': f'{updates} atualizações processadas.'})


def roster_calendar_feed(request, signed_token):
    """
    Gera um feed iCal (.ics) criptograficamente assinado com as escalas personalizadas
    (ShiftOverride) e as licenças aprovadas do funcionário.
    """
    from django.core.signing import Signer, BadSignature
    from django.http import HttpResponse, HttpResponseForbidden
    from django.shortcuts import get_object_or_404
    from leave.models import Leave
    from datetime import timedelta
    
    signer = Signer()
    try:
        employee_id = signer.unsign(signed_token)
    except BadSignature:
        return HttpResponseForbidden("Assinatura de link inválida ou expirada.")

    employee = get_object_or_404(Employee, pk=employee_id)

    # Definir faixa de exportação: 60 dias no passado até 120 dias no futuro
    today = timezone.localdate()
    start_date = today - timedelta(days=60)
    end_date = today + timedelta(days=120)

    # Escalas personalizadas (exceções de trabalho e folgas)
    overrides = ShiftOverride.objects.filter(
        employee=employee,
        date__range=[start_date, end_date]
    )

    # Licenças aprovadas
    leaves = Leave.objects.filter(
        employee=employee,
        status='APPROVED',
        date__range=[start_date, end_date]
    ).select_related('leave_type')

    # Monta o arquivo iCalendar (RFC 5545)
    lines = []
    lines.append("BEGIN:VCALENDAR")
    lines.append("VERSION:2.0")
    lines.append("PRODID:-//Netline RH//Roster Feed//PT")
    lines.append("CALSCALE:GREGORIAN")
    lines.append("METHOD:PUBLISH")
    lines.append(f"X-WR-CALNAME:Netline RH - {employee.full_name}")
    lines.append("X-WR-TIMEZONE:America/Sao_Paulo")

    # 1. Adicionar exceções de turno
    for ov in overrides:
        uid = f"override-{ov.id}-{employee.id}@netlinerh.com.br"
        stamp = timezone.now().strftime("%Y%m%dT%H%M%SZ")
        
        if ov.override_type == ShiftOverride.TYPE_WORK:
            if ov.entry_time and ov.exit_time:
                from datetime import datetime
                dt_start = datetime.combine(ov.date, ov.entry_time)
                
                # Tratar plantões noturnos (horário de saída anterior ou igual ao de entrada)
                if ov.exit_time <= ov.entry_time:
                    dt_end = datetime.combine(ov.date + timedelta(days=1), ov.exit_time)
                else:
                    dt_end = datetime.combine(ov.date, ov.exit_time)
                
                # Converter datas ingênuas (naive) para cientes do fuso horário atual (America/Sao_Paulo) e depois para UTC
                if timezone.is_naive(dt_start):
                    dt_start = timezone.make_aware(dt_start, timezone.get_current_timezone())
                if timezone.is_naive(dt_end):
                    dt_end = timezone.make_aware(dt_end, timezone.get_current_timezone())
                
                from datetime import timezone as datetime_timezone
                dt_start_utc = dt_start.astimezone(datetime_timezone.utc)
                dt_end_utc = dt_end.astimezone(datetime_timezone.utc)
                
                start_str = dt_start_utc.strftime("%Y%m%dT%H%M%SZ")
                end_str = dt_end_utc.strftime("%Y%m%dT%H%M%SZ")
                
                summary = f"Turno: {ov.entry_time.strftime('%H:%M')} - {ov.exit_time.strftime('%H:%M')}"
                if ov.reason:
                    display_reason = ov.reason.replace("Escala Viva - ", "").replace("Turno via Excel", "Turno")
                    summary = f"Turno: {display_reason} ({ov.entry_time.strftime('%H:%M')}-{ov.exit_time.strftime('%H:%M')})"
                
                lines.append("BEGIN:VEVENT")
                lines.append(f"UID:{uid}")
                lines.append(f"DTSTAMP:{stamp}")
                lines.append(f"DTSTART:{start_str}")
                lines.append(f"DTEND:{end_str}")
                lines.append(f"SUMMARY:{summary}")
                lines.append(f"DESCRIPTION:Escala de trabalho agendada no Netline RH. Turno: {ov.entry_time.strftime('%H:%M')} ate {ov.exit_time.strftime('%H:%M')}. Motivo: {ov.reason or 'Nao informado'}")
                lines.append("STATUS:CONFIRMED")
                lines.append("END:VEVENT")
        
        elif ov.override_type == ShiftOverride.TYPE_REST:
            start_str = ov.date.strftime("%Y%m%d")
            end_str = (ov.date + timedelta(days=1)).strftime("%Y%m%d")
            
            lines.append("BEGIN:VEVENT")
            lines.append(f"UID:{uid}")
            lines.append(f"DTSTAMP:{stamp}")
            lines.append(f"DTSTART;VALUE=DATE:{start_str}")
            lines.append(f"DTEND;VALUE=DATE:{end_str}")
            lines.append("SUMMARY:Folga / DSR")
            lines.append(f"DESCRIPTION:Dia de descanso programado. Motivo: {ov.reason or 'Folga / DSR'}")
            lines.append("STATUS:CONFIRMED")
            lines.append("END:VEVENT")

    # 2. Adicionar licenças aprovadas
    for lv in leaves:
        uid = f"leave-{lv.id}-{employee.id}@netlinerh.com.br"
        stamp = timezone.now().strftime("%Y%m%dT%H%M%SZ")
        
        start_str = lv.date.strftime("%Y%m%d")
        end_str = (lv.date + timedelta(days=1)).strftime("%Y%m%d")
        
        desc = f"Licenca / Afastamento aprovado. Tipo: {lv.leave_type.name}."
        if lv.duration_type == Leave.LENGTH_HALF:
            desc += " Periodo: Meio Periodo."
            summary = f"Licenca (Meio Periodo): {lv.leave_type.name}"
        else:
            summary = f"Licenca: {lv.leave_type.name}"
            
        lines.append("BEGIN:VEVENT")
        lines.append(f"UID:{uid}")
        lines.append(f"DTSTAMP:{stamp}")
        lines.append(f"DTSTART;VALUE=DATE:{start_str}")
        lines.append(f"DTEND;VALUE=DATE:{end_str}")
        lines.append(f"SUMMARY:{summary}")
        lines.append(f"DESCRIPTION:{desc}")
        lines.append("STATUS:CONFIRMED")
        lines.append("END:VEVENT")

    lines.append("END:VCALENDAR")
    
    ical_content = "\r\n".join(lines)
    
    response = HttpResponse(ical_content, content_type="text/calendar; charset=utf-8")
    response['Content-Disposition'] = f'attachment; filename="escala_{employee.id}.ics"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

